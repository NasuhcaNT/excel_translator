import os
import sys
import pandas as pd
from openpyxl import load_workbook
from deep_translator import GoogleTranslator


def translate_excel(input_file, output_file):
    """Translates text in an Excel file from English to Turkish while preserving formatting."""
    if not os.path.exists(input_file):
        print(f"Error: {input_file} not found!")
        return

    # Load the Excel file
    wb = load_workbook(input_file)
    translator = GoogleTranslator(source="en", target="tr")

    new_wb = load_workbook(input_file)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        new_sheet = new_wb[sheet_name]

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    translated_text = translator.translate(cell.value)
                    new_sheet[cell.coordinate] = translated_text

    new_wb.save(output_file)
    print(f"Translation completed! New file saved: {output_file}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python translate.py <input_file.xlsx> <output_file.xlsx>")
    else:
        input_excel = sys.argv[1]
        output_excel = sys.argv[2]
        print(f"Starting translation: {input_excel} -> {output_excel}")
        translate_excel(input_excel, output_excel)
